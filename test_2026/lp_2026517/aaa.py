import polars as pl

def read(path_self: str):
    """
    从 Excel 中提取已知月份、已知数值、缺失月份。
    要求 Excel 结构：
        - 第一行：月份数字（如 1,2,3,4,5,6）
        - 第二行：对应数值（如 10,11,13,空,20,25）
    返回三个列表：
        x_known, y_known, missing_months
    """
    # 读取 Excel，不将第一行作为表头（has_header=False）
    df = pl.read_excel(path_self, has_header=False)

    # 至少需要两行数据
    if df.height < 2:
        raise ValueError("Excel 中至少需要两行：月份行和数值行。")

    month_row = df.row(0)   # 第一行：月份
    value_row = df.row(1)   # 第二行：数值

    x_known = []
    y_known = []
    missing_months = []

    for m, v in zip(month_row, value_row):
        month_int = int(m)          # 确保月份是整数
        if v is not None:
            x_known.append(month_int)
            y_known.append(v)
        else:
            missing_months.append(month_int)

    return x_known, y_known, missing_months


import openpyxl

def write(file_path: str, months: list, values: list):
    """
    将完整的月份和数值写回 Excel 文件（覆盖第一个工作表）。
    格式：第一行月份数字，第二行对应数值。
    使用 openpyxl，无需额外安装 xlsxwriter 或 fastexcel。
    """
    # 尝试打开已有文件（如果存在），否则创建新工作簿
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 使用第一个工作表，清空原有内容
    ws = wb.active
    ws.title = "Sheet1"
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)

    # 写入第一行：月份数字
    for col_idx, month in enumerate(months, start=1):
        ws.cell(row=1, column=col_idx, value=month)

    # 写入第二行：数值
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # 自动调整列宽（简单实现）
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # 保存文件
    wb.save(file_path)
# ========== 主程序 ==========


def write(months: list, values: list, output_path: str):
    """
    将补全后的月份和数值写入 Excel 文件。
    保持与 read 函数一致的格式：第一行为月份，第二行为数值（横向排列）。
    """
    # 构建字典，将月份和数值打包成两行数据（注意这里用列表包列表）
    data = {
        "row1": [months],   # 第一行：月份列表
        "row2": [values]    # 第二行：数值列表
    }
    
    # 创建 Polars DataFrame
    df_out = pl.DataFrame(data)
    
    # 写入 Excel
    # write_schema=False 防止把列名(row1, row2)写进 Excel
    # header=False 防止写入额外的表头，完美还原纯数据格式
    df_out.write_excel(output_path, write_schema=False, header=False, autofit=True)
    print(f"✅ 数据已成功保存至: {output_path}")
    
    
# ========== 主程序 ==========
if __name__ == "__main__":
    path_self = "temp/test(1).xlsx"
    
    try:
        # 1. 读取原始数据
        x_known, y_known, missing = read(path_self)

        print(f"已知月份: {x_known}")
        print(f"已知数值: {y_known}")
        print(f"缺失月份: {missing}")
        
        # 2. 模拟插值过程
        all_months = sorted(x_known + missing)
        all_values = []
        
        # 将已知数据转为字典，提高后续查找效率（避免在循环中反复使用 index 查找）
        known_dict = dict(zip(x_known, y_known))
        avg_val = sum(y_known) / len(y_known) # 模拟插值的均值
        
        for m in all_months:
            if m in known_dict:
                val = known_dict[m]  # 已知月份，直接取值
            else:
                val = avg_val        # 缺失月份，填入模拟的插值结果
            all_values.append(val)
        
        # 3. 覆盖写入原表
        write(all_months, all_values, path_self)
    except Exception as e:
        print(f"❌ 发生错误: {e}")